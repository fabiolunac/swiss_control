from openpyxl import load_workbook
from datetime import datetime
import tkinter as tk 
from parameters import parse_date_input

def add_line(path, dados, sheet_name=None):
    """
    Adiciona uma linha (lista/tupla) ao final da planilha.
    """
    wb = load_workbook(path)

    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    ws.append(dados)
    ws.cell(row=ws.max_row, column=1).number_format = "DD/MM/YYYY"
    wb.save(path)

def main():
    path = "/Users/fabioluna/OneDrive - CERN/Swiss Control.xlsx"

    today = datetime.now().strftime('%d/%m')

    data = input(f"Data (DD/MM) [{datetime.now().strftime('%d/%m')}]: ").strip()
    if not data:
        data = today

    try:
        data = parse_date_input(data)
    except ValueError as exc:
        print(str(exc))
        return

    local = input("Local: ").strip()
    valor_raw = input("Valor: ").strip()

    try:
        valor = float(valor_raw.replace(",", "."))
    except ValueError:
        valor = valor_raw

    dados = [data, local, valor]

    print(dados)
    add_line(path, dados)

if __name__ == "__main__":
    main()
